import os
import re
import io
import math
import traceback
from typing import List, Optional, Tuple
import tempfile
from datetime import datetime

import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font

# ==================== 설정 ====================
st.set_page_config(
    page_title="Excursion App V7",
    page_icon="🎵",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== 상수 ====================
APP_TITLE = "Excursion App V7"

AMP_CONFIG = {
    "Goodix Top":    {"col_letter": "Q", "start_row": 10},
    "Goodix Bottom": {"col_letter": "R", "start_row": 10},
    "TI Top":        {"col_letter": "H", "start_row": 11},
    "TI Bottom":     {"col_letter": "P", "start_row": 11},
    "Awinic Top":    {"col_letter": "B", "start_row": 3},
    "Awinic Bottom": {"col_letter": "C", "start_row": 3},
}

# ==================== 유틸리티 함수 ====================
def safe_float(value) -> Optional[float]:
    """문자열/숫자를 float로 안전하게 변환"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            if math.isnan(value):
                return None
        except Exception:
            pass
        return float(value)

    s = str(value).replace("\ufeff", "").replace("\x00", "").strip()
    s = s.strip('"').strip("'").strip()

    if not re.fullmatch(r"[-+]?(?:\d+(?:\.\d*)?|\.\d+)(?:[eE][-+]?\d+)?", s):
        return None

    try:
        return float(s)
    except Exception:
        return None


def flatten_excel_values(values):
    """Excel 값 플래튼"""
    if values is None:
        return []
    if not isinstance(values, tuple):
        return [values]

    out = []
    for item in values:
        if isinstance(item, tuple):
            out.append(item[0] if len(item) else None)
        else:
            out.append(item)
    return out


def read_column_with_excel_engine(path: str, col_letter: str, start_row: int) -> Tuple[List[float], str]:
    """Microsoft Excel 엔진으로 열 읽기"""
    try:
        import win32com.client
    except Exception as e:
        raise RuntimeError(
            "❌ Excel 엔진 사용을 위해 pywin32가 필요합니다.\n\n"
            "pip install pywin32"
        ) from e

    excel = None
    wb = None

    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(os.path.abspath(path))
        ws = wb.Worksheets(1)
        used = ws.UsedRange
        last_row = used.Row + used.Rows.Count - 1

        if last_row < start_row:
            raise ValueError(f"데이터 마지막 행이 시작행보다 작습니다. start={start_row}, last={last_row}")

        rng = ws.Range(f"{col_letter}{start_row}:{col_letter}{last_row}")
        raw_values = flatten_excel_values(rng.Value)

        values_all = [safe_float(v) for v in raw_values]
        numeric_values = [v for v in values_all if v is not None]
        start_idx = next((i for i, v in enumerate(values_all) if v is not None and v > 0.01), None)

        debug = (
            f"Excel COM 엔진 읽음\n"
            f"파일={os.path.basename(path)}\n"
            f"열={col_letter}, start_row={start_row}, last_row={last_row}\n"
            f"Raw count={len(raw_values)}, Numeric count={len(numeric_values)}, "
            f"Max={max(numeric_values) if numeric_values else 'N/A'}"
        )

        if not numeric_values:
            raise ValueError("대상 열에서 숫자를 찾지 못했습니다.\n\n" + debug)
        if start_idx is None:
            raise ValueError("대상 열에 0.01 초과 데이터가 없습니다.\n\n" + debug)

        final = [v for v in values_all[start_idx:] if v is not None]
        if not final:
            raise ValueError("데이터 추출 결과가 비어 있습니다.\n\n" + debug)

        return final, debug

    finally:
        try:
            if wb is not None:
                wb.Close(False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass


def read_column_from_csv(path: str, col_letter: str, start_row: int) -> Tuple[List[float], str]:
    """CSV 파일에서 열 읽기"""
    try:
        col_idx = ord(col_letter) - ord('A')
        
        delimiters = [',', '\t', ';', '|', ' ']
        df = None
        
        for delim in delimiters:
            try:
                df = pd.read_csv(
                    path,
                    sep=delim,
                    header=None,
                    engine='python',
                    on_bad_lines='skip',
                    skiprows=start_row - 1
                )
                if len(df.columns) > col_idx:
                    break
            except Exception:
                continue
        
        if df is None or len(df.columns) <= col_idx:
            raise ValueError(f"CSV 파일에 열 '{col_letter}' (인덱스 {col_idx})가 없습니다.")
        
        column_data = df.iloc[:, col_idx].tolist()
        
        values_all = [safe_float(v) for v in column_data]
        numeric_values = [v for v in values_all if v is not None]
        start_idx = next((i for i, v in enumerate(values_all) if v is not None and v > 0.01), None)

        debug = (
            f"CSV 파일 읽음\n"
            f"파일={os.path.basename(path)}\n"
            f"열={col_letter} (인덱스 {col_idx}), start_row={start_row}"
        )

        if not numeric_values:
            raise ValueError("대상 열에서 숫자를 찾지 못했습니다.\n\n" + debug)
        if start_idx is None:
            raise ValueError("대상 열에 0.01 초과 데이터가 없습니다.\n\n" + debug)

        final = [v for v in values_all[start_idx:] if v is not None]
        if not final:
            raise ValueError("데이터 추출 결과가 비어 있습니다.\n\n" + debug)

        return final, debug

    except Exception as e:
        raise Exception(f"CSV 파일 읽기 실패: {str(e)}")


def read_logging_column(file_bytes: bytes, filename: str, amp: str) -> Tuple[List[float], str]:
    """로깅 파일에서 AMP별 데이터 읽기"""
    cfg = AMP_CONFIG[amp]
    
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        if filename.lower().endswith('.csv'):
            return read_column_from_csv(tmp_path, cfg["col_letter"], cfg["start_row"])
        else:
            return read_column_with_excel_engine(tmp_path, cfg["col_letter"], cfg["start_row"])
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def read_real_column(file_bytes: bytes, filename: str) -> List[float]:
    """실측 파일에서 C열 데이터 읽기"""
    with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(filename)[1]) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name

    try:
        if filename.lower().endswith('.csv'):
            values_all, _ = read_column_from_csv(tmp_path, "C", 1)
        else:
            values_all, _ = read_column_with_excel_engine(tmp_path, "C", 1)

        interval = 500
        processed = []
        for i in range(0, len(values_all), interval):
            seg = values_all[i:i + interval]
            if seg:
                processed.append(max(abs(x) for x in seg))
        return processed
    finally:
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass


def pad_list(values: List[float], length: int):
    """리스트 패딩"""
    return values + [math.nan] * (length - len(values))


def max_omit_nan(values):
    """NaN 제외 최댓값"""
    vals = [v for v in values if v is not None and not (isinstance(v, float) and math.isnan(v))]
    return max(vals) if vals else math.nan


def mean_omit_nan(values):
    """NaN 제외 평균"""
    vals = [v for v in values if v is not None and not (isinstance(v, float) and math.isnan(v))]
    return sum(vals) / len(vals) if vals else math.nan


def calculate_result(log_bytes: bytes, log_filename: str, real_bytes: bytes, real_filename: str, amp: str):
    """결과 계산"""
    if not log_bytes:
        raise ValueError("로깅파일을 선택하세요.")
    if not real_bytes:
        raise ValueError("실측파일을 선택하세요.")

    log_data, log_debug = read_logging_column(log_bytes, log_filename, amp)
    real_data = read_real_column(real_bytes, real_filename)

    max_len = max(len(log_data), len(real_data))
    if max_len == 0:
        raise ValueError("로깅/실측 데이터가 모두 비어 있습니다.")

    final_log = pad_list(log_data, max_len)
    final_real = pad_list(real_data, max_len)

    num_group_300 = math.ceil(max_len / 3)
    logging_300_temp = []
    real_300_temp = []

    for i in range(num_group_300):
        s = i * 3
        e = min((i + 1) * 3, max_len)
        logging_300_temp.append(max_omit_nan(final_log[s:e]))
        real_300_temp.append(max_omit_nan(final_real[s:e]))

    logging_300 = pad_list(logging_300_temp, max_len)
    real_300 = pad_list(real_300_temp, max_len)

    compare_len = min(65, num_group_300)
    consistency_diff = [math.nan] * max_len

    for i in range(compare_len):
        a = real_300[i]
        b = logging_300[i]
        if not math.isnan(a) and not math.isnan(b):
            consistency_diff[i] = abs(a - b)

    excursion_consistency_value = mean_omit_nan(consistency_diff[:compare_len])
    excursion_real_max_value = max_omit_nan(final_real)

    # DataFrame 생성
    out_df = pd.DataFrame({
        "Logging_Data_100ms": final_log,
        "Real_Data_100ms": final_real,
        "Logging_Data_300ms": logging_300,
        "Real_Data_300ms": real_300,
        "Real - Logging Delta": consistency_diff,
    })

    return {
        "df": out_df,
        "consistency": excursion_consistency_value,
        "real_max": excursion_real_max_value,
        "log_data": final_log[:max_len],
        "real_data": final_real[:max_len],
        "max_len": max_len,
    }


def calculate_consistency_with_step(log_data: List[float], real_data: List[float], log_step: float, real_step: float):
    """X step을 고려하여 정합성 재계산 (기본값 0.1s 기준)"""
    # 기본 샘플 간격: 100ms = 0.1s
    base_step = 0.1
    
    # X step에 따른 샘플 간격 계산
    log_interval = max(1, round(log_step / base_step))
    real_interval = max(1, round(real_step / base_step))
    
    # 인덱싱으로 데이터 선택
    resampled_log = [log_data[i] for i in range(0, len(log_data), log_interval)]
    resampled_real = [real_data[i] for i in range(0, len(real_data), real_interval)]
    
    # 길이 맞추기
    max_len = max(len(resampled_log), len(resampled_real))
    final_log = pad_list(resampled_log, max_len)
    final_real = pad_list(resampled_real, max_len)
    
    # 300ms 단위로 묶기
    num_group_300 = math.ceil(max_len / 3)
    logging_300_temp = []
    real_300_temp = []

    for i in range(num_group_300):
        s = i * 3
        e = min((i + 1) * 3, max_len)
        logging_300_temp.append(max_omit_nan(final_log[s:e]))
        real_300_temp.append(max_omit_nan(final_real[s:e]))

    logging_300 = pad_list(logging_300_temp, max_len)
    real_300 = pad_list(real_300_temp, max_len)

    # 정합성 계산
    compare_len = min(65, num_group_300)
    consistency_diff = [math.nan] * max_len

    for i in range(compare_len):
        a = real_300[i]
        b = logging_300[i]
        if not math.isnan(a) and not math.isnan(b):
            consistency_diff[i] = abs(a - b)

    new_consistency = mean_omit_nan(consistency_diff[:compare_len])
    new_real_max = max_omit_nan(final_real)
    
    return new_consistency, new_real_max


def save_result_excel(result, amp: str, xmax_value: float = math.nan):
    """결과를 Excel로 저장"""
    try:
        out_df = result["df"]
        max_len = result["max_len"]
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Excursion_{amp}_{timestamp}.xlsx"
        save_path = os.path.join(os.getcwd(), filename)
        
        out_df.to_excel(save_path, index=False, sheet_name="Data")
        
        wb = load_workbook(save_path)
        ws = wb.active
        
        ws["I1"] = "Result"
        ws["I2"] = "Excursion 정합성 (mm)"
        ws["J2"] = result["consistency"]
        ws["I3"] = "Excursion Real Max (mm)"
        ws["J3"] = result["real_max"]
        ws["I4"] = "Speaker Xmax (mm)"
        ws["J4"] = xmax_value if not (isinstance(xmax_value, float) and math.isnan(xmax_value)) else None
        
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        dark_font = Font(color="003B73", bold=True)
        for cell in ["I1", "I2", "I3", "I4"]:
            ws[cell].fill = header_fill
            ws[cell].font = dark_font
        
        chart = LineChart()
        chart.title = f"Processed Data - {amp} (Top 700 points)"
        chart.y_axis.title = "Excursion (mm)"
        chart.x_axis.title = "Index"
        last_row = min(max_len + 1, 701)
        data = Reference(ws, min_col=1, max_col=2, min_row=1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        
        if not (isinstance(xmax_value, float) and math.isnan(xmax_value)):
            xmax_col = 8
            ws.cell(row=1, column=xmax_col, value="Speaker Xmax")
            for r in range(2, last_row + 1):
                ws.cell(row=r, column=xmax_col, value=xmax_value)
            xmax_ref = Reference(ws, min_col=xmax_col, max_col=xmax_col, min_row=1, max_row=last_row)
            chart.add_data(xmax_ref, titles_from_data=True)
        
        ws.add_chart(chart, "I7")
        wb.save(save_path)
        
        return save_path
    except Exception as e:
        raise Exception(f"Excel 저장 실패: {str(e)}")


# ==================== Streamlit 앱 ====================
def main():
    st.title("🎵 " + APP_TITLE)
    
    if 'current_result' not in st.session_state:
        st.session_state.current_result = None
    if 'log_step' not in st.session_state:
        st.session_state.log_step = 0.1
    if 'real_step' not in st.session_state:
        st.session_state.real_step = 0.1
    
    # 사이드바
    with st.sidebar:
        st.header("⚙️ 설정")
        
        selected_amp = st.selectbox("AMP 선택", list(AMP_CONFIG.keys()), index=0)
        
        st.divider()
        st.subheader("📊 그래프 축 설정")
        
        log_step = st.number_input(
            "Logging X step [s]",
            value=st.session_state.log_step,
            min_value=0.0001,
            step=0.0001,
            format="%.4f",
            key="log_step_input"
        )
        st.session_state.log_step = log_step
        
        real_step = st.number_input(
            "Real X step [s]",
            value=st.session_state.real_step,
            min_value=0.0001,
            step=0.0001,
            format="%.4f",
            key="real_step_input"
        )
        st.session_state.real_step = real_step
        
        st.divider()
        xmax = st.number_input("Speaker Xmax (mm)", value=0.0, step=0.001, format="%.4f")
    
    # 메인
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.subheader("📁 입력 파일")
        col_log, col_real = st.columns(2)
        
        with col_log:
            st.markdown("**로깅파일** 📥")
            log_file = st.file_uploader("로깅 파일 (Excel/CSV)", type=["xlsx", "xls", "xlsm", "csv"], key="log_file")
        
        with col_real:
            st.markdown("**실측파일** 📥")
            real_file = st.file_uploader("실측 파일 (Excel/CSV)", type=["xlsx", "xls", "xlsm", "csv"], key="real_file")
    
    with col2:
        st.subheader("📊 결과")
        result_placeholder = st.empty()
    
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        run_button = st.button("▶️ 실행", use_container_width=True, key="run_btn")
    
    with col2:
        reset_button = st.button("🔄 초기화", use_container_width=True, key="reset_btn")
    
    if reset_button:
        st.session_state.current_result = None
        st.rerun()
    
    if run_button:
        if not log_file or not real_file:
            st.error("❌ 로깅파일과 실측파일을 모두 선택하세요!")
        else:
            try:
                with st.spinner("⏳ 처리 중... 파일을 읽는 중입니다."):
                    result = calculate_result(log_file.getvalue(), log_file.name, real_file.getvalue(), real_file.name, selected_amp)
                    st.session_state.current_result = result
                    st.session_state.current_amp = selected_amp
                    st.session_state.current_xmax = xmax
                    st.success("✅ 처리 완료!")
            except Exception as e:
                st.error(f"❌ 오류 발생: {str(e)}")
                with st.expander("상세 오류 정보"):
                    st.code(traceback.format_exc())
    
    if st.session_state.current_result:
        result = st.session_state.current_result
        current_amp = st.session_state.get('current_amp', selected_amp)
        current_xmax = st.session_state.get('current_xmax', xmax)
        
        # X step에 따라 정합성 재계산 ⭐ 핵심 로직
        consistency_val, real_max_val = calculate_consistency_with_step(
            result["log_data"], 
            result["real_data"], 
            log_step, 
            real_step
        )
        
        # 결과 표시
        with col2:
            with result_placeholder.container():
                col_metric1, col_metric2 = st.columns(2)
                with col_metric1:
                    st.metric(
                        "Excursion 정합성",
                        f"{consistency_val:.3f} mm" if not math.isnan(consistency_val) else "NaN"
                    )
                with col_metric2:
                    st.metric(
                        "Real Max",
                        f"{real_max_val:.3f} mm" if not math.isnan(real_max_val) else "NaN"
                    )
        
        st.divider()
        st.subheader("📈 그래프")
        
        fig, ax = plt.subplots(figsize=(12, 5))
        fig.patch.set_facecolor('white')
        ax.set_facecolor('white')
        
        log_data = result["log_data"]
        real_data = result["real_data"]
        n_log = min(len(log_data), 700)
        n_real = min(len(real_data), 700)
        
        x_log = [i * log_step for i in range(n_log)]
        x_real = [i * real_step for i in range(n_real)]
        
        ax.plot(x_log, log_data[:n_log], marker=".", markersize=0.8, linewidth=2.0, label="Logging_Data_100ms", color="#003B73")
        ax.plot(x_real, real_data[:n_real], marker=".", markersize=0.8, linewidth=2.0, label="Real_Data_100ms", color="#FF6B6B")
        
        if current_xmax > 0:
            ax.axhline(y=current_xmax, color="red", linestyle="--", linewidth=1.6, label="Speaker Xmax")
        
        ax.set_title(f"Processed Data - {current_amp}", fontsize=14, fontweight='bold', color="#003B73")
        ax.set_xlabel("Time [s]", fontsize=12, color="#000000")
        ax.set_ylabel("Excursion (mm)", fontsize=12, color="#000000")
        ax.tick_params(colors="#000000")
        ax.grid(True, alpha=0.3, color="#CCCCCC")
        
        for spine in ax.spines.values():
            spine.set_color("#000000")
        
        ax.legend(fontsize=10)
        fig.tight_layout()
        st.pyplot(fig)
        
        st.divider()
        st.subheader("📋 데이터 통계")
        
        col_stat1, col_stat2, col_stat3 = st.columns(3)
        with col_stat1:
            st.info(f"**Logging 데이터 개수**\n{len(log_data)}")
        with col_stat2:
            st.info(f"**Real 데이터 개수**\n{len(real_data)}")
        with col_stat3:
            if log_data:
                first_val = log_data[0]
                st.info(f"**Logging 첫 값**\n{first_val:.4f}" if not math.isnan(first_val) else "**Logging 첫 값**\nNaN")
        
        # 저장 버튼
        st.divider()
        col_save1, col_save2 = st.columns([1, 3])
        
        with col_save1:
            if st.button("💾 Excel 저장", use_container_width=True):
                try:
                    with st.spinner("📊 Excel 파일을 생성 중입니다..."):
                        save_path = save_result_excel(result, current_amp, current_xmax)
                        st.success(f"✅ 저장 완료!\n{save_path}")
                except Exception as e:
                    st.error(f"❌ 저장 실패: {str(e)}")


if __name__ == "__main__":
    main()